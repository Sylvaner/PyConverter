#!/usr/bin/env python3
"""
Convert and transform table file to another one.
Create config file for conversion informations (see parse_config)
"""
__author__ = "Sylvain Dangin"
__licence__ = "Apache 2.0"
__version__ = "1.0"
__maintainer__ = "Sylvain Dangin"
__email__ = "sylvain.dangin@gmail.com"
__status__ = "Development"

import io
import os
import sys
import csv
import json
import openpyxl
import xlrd
import xlwt

class Convert():
    # Unknown file type
    UNKNOWN_FILE = 0
    # Old excel file type (xls)
    OLD_EXCEL_FILE = 1
    # Excel file type (xlsx)
    NEW_EXCEL_FILE = 2
    # CSV file type
    CSV_FILE = 3
    # List of know extesions
    EXT_LIST = {}
    # Folder for transforms modules
    TRANSFORMS_FOLDER = 'transforms'
    # Folder for actions modules a
    ACTIONS_FOLDER = 'actions'

    # List of modules loaded dynamically
    dynamics_modules = {}
    # Type of input file
    input_file_type = None
    # Type of output file
    output_file_type = None
    # First line of input file is the name of the columns
    input_first_line_header = True
    # Remove header if present in input file
    ignore_first_line_header = False
    # CSV columns delimiter for input file
    input_csv_delimiter = ';'
    # CSV columns delimiter for output file
    output_csv_delimiter = ';'
    # Name of the sheet for Excel import
    input_xls_sheet_name = None
    # Input encoding for CSV file
    input_encoding = ''
    # Ouput encoding for CSV file
    output_encoding = ''
    # Name of the sheet for Excel export
    output_xls_sheet_name = 'export'
    # Moves of columns
    moves = []
    # Number of columns in the ouput file
    output_number_of_cols = 0
    # Header of input columns
    input_header = None
    # Data of the input file
    input_data = None
    # Data of the output file
    output_data = None
    
    def __init__(self):
        """Constructor
            - Initialize some default values.
            - Add transforms and actions folders
        """
        self.EXT_LIST['csv'] = self.CSV_FILE
        self.EXT_LIST['xls'] = self.OLD_EXCEL_FILE
        self.EXT_LIST['xlsx'] = self.NEW_EXCEL_FILE
        if os.path.exists(self.TRANSFORMS_FOLDER):
            sys.path.append(self.TRANSFORMS_FOLDER)
        if os.path.exists(self.ACTIONS_FOLDER):
            sys.path.append(self.ACTIONS_FOLDER)

    def set_default_values(self):
        """Set all default values of the class
        """
        self.input_file_type = self.UNKNOWN_FILE
        self.output_file_type = self.UNKNOWN_FILE
        # Set default encoding depending of the OS
        if 'nt' in os.name:
            self.input_encoding = 'iso-8859-1'
            self.output_encoding = 'iso-8859-1'
        else:
            self.input_encoding = 'utf-8'
            self.output_encoding = 'utf-8'
        self.input_first_line_header = True
        self.ignore_first_line_header = False
        self.input_csv_delimiter = ';'
        self.output_csv_delimiter = ';'
        self.input_xls_sheet_name = None
        self.output_xls_sheet_name = 'export'
        self.moves = []
        self.output_number_of_cols = 0
        self.input_header = None
        self.input_data = None
        self.output_data = None
        
    def get_file_type_by_ext(self, filename):
        """Determine extension from the filename.

        :param filename: Path of the input file

        :return int:
         - Convert.UNKNOWN_FILE (0) for unknown file,
         - Convert.OLD_EXCEL_FILE (1) for Excel file,
         - Convert.NEW_EXCEL_FILE (2) for Excel file,
         - Convert.CSV_FILE (3) for csv file.
        """
        file_type = self.UNKNOWN_FILE

        # Test all knowed files extensions
        for ext, ext_type in self.EXT_LIST.items():
            if '.' in filename and ext in filename:
                # Extract extension
                filename_parts = filename.split('.')
                ext_part = filename_parts[len(filename_parts) -1]
                if ext_part == ext:
                    file_type = ext_type
        return file_type

    def read_config(self, config):
        """Read configuration informations from differents formats

        :param config: Configuration informations
         - dict: Dictionnary with informations,
         - str:
           - json string: JSON valid string,
           - path: Path to the configuration file,
         - file: Opened file.
        """
        json_config = None
        # Config already at json format
        if isinstance(config, dict):
            json_config = config
        # Opened file
        elif isinstance(config, io.IOBase):
            try:
                json_config = json.load(config)
            except json.decoder.JSONDecodeError as e:
                print(e)
                raise Exception('Error: Config file bad format.')
        # String
        elif isinstance(config, str):
            # Test if string is json object
            try:
                json_config = json.loads(config)
            except json.decoder.JSONDecodeError as e:
                # Test if string if filename
                if os.path.exists(config):
                    try:
                        # Open as json
                        config_file = open(config, 'r')
                        json_config = json.load(config_file)
                        config_file.close()
                    except FileNotFoundError as fnf_e:
                        raise Exception('Error: Unable to read config file.')
                    except json.decoder.JSONDecodeError as e:
                        raise Exception('Error: Config file bad format')
                else:
                    raise Exception('Error: Unable to read config')
        self.parse_config(json_config)

    def parse_config(self, config):
        """Parse parameters from configuration.
        
        :param config: Configuration informations in JSON format 
                       from string or file.

        List of JSON parameters:
         - input_file_type: Format of the input file.
         - output_file_type: Format of the output file.
         - input_csv_delimiter: CSV column delimiter (Default: ;)
         - output_csv_delimiter: CSV column delimiter (Default: ;)
         - input_first_line_header: First line of input file 
           (Default: True)
         - ignore_first_line_header: Remove first line header in output 
           (Default: False)
         - input_xls_sheet_name: Title of the sheet for Excel input 
           (Default: first worksheet)
         - output_xls_sheet_name: Title of the sheet for Excel output 
           (Default: export)
         - input_encoding : Encoding using in CSV input file
         - output_encoding : Encoding using in CSV output file
         - moves: List of columns moves
            Examples :
              1) Switch column 1 and 2
                 {
                   "moves": [
                     {"from": 1, "to": 2}, 
                     {"from": 2, "to": 1}
                   ]
                 }
              2) Move column 3 to 1 and remove right and left spaces
                 {"moves": {"from": 3, "to": 1, "transform": "trim"}}
              3) Keep column 2 and set column 1 to TEST : 
                 {
                   "moves":
                   [
                     {"from": 2, "to": 2}, 
                     {"from": 1, "to": 1, "action": {"set_value": "TEST"}
                   ]
                 }
        """

        # Read simple config informations
        attr_list = ['input_csv_delimiter',
                     'output_csv_delimiter',
                     'input_first_line_header',
                     'ignore_first_line_header',
                     'output_xls_sheet_name',
                     'input_encoding',
                     'output_encoding']
        for attr in attr_list:
            if attr in config.keys():
                setattr(self, attr, config[attr])
                
        if 'input_file_type' in config.keys():
            self.input_file_type = self.EXT_LIST[config['input_file_type']]
        if 'output_file_type' in config.keys():
            self.output_file_type = self.EXT_LIST[config['output_file_type']]

        if 'moves' in config:
            # List of moves
            if isinstance(config['moves'], dict):
                self.moves = [config['moves']]
            elif isinstance(config['moves'], list):
                self.moves = config['moves']
            max_col = 0
            # Get number of columns in output file
            for move in self.moves:
                # Check transform library
                if 'transform' in move:
                    self.load_modules(self.TRANSFORMS_FOLDER, move['transform'])
                if 'action' in move:
                    self.load_modules(self.ACTIONS_FOLDER, move['action'])
                dest = int(move['to'])
                if dest > max_col:
                    max_col = dest
            self.output_number_of_cols = max_col

    def load_modules(self, folder, modules_data):
        """Load modules used by the config
        :param folder: Folder of modules
        :param module_data: Informations about the module to load
        """
        # Read list of modules
        if isinstance(modules_data, list):
            for module_item in modules_data:
                self.load_modules(folder, module_item)
        # Read module name from dict
        elif isinstance(modules_data, dict):
            for module_item in modules_data:
                self.load_modules(folder, module_item)
        # Test if module file exists
        elif not os.path.exists(folder+os.path.sep+modules_data+'.py'):
            raise Exception('Error: missing transform "'+modules_data+'"')
        # If module not already loaded, load it
        elif not modules_data in self.dynamics_modules:
            self.dynamics_modules[modules_data] = getattr(__import__(modules_data), modules_data)

    def read_input_data_from_csv(self, input_filename):
        """Read input data from CSV file
        :param input_filename: Name of input file
        """
        self.input_header = []
        self.input_data = []

        if not os.path.exists(input_filename):
            raise Exception('Error: Input file not found.')

        with open(input_filename, 'r', encoding=self.input_encoding) as input_file:
            reader = csv.reader(input_file, delimiter = self.input_csv_delimiter)
            for row in reader:
                if reader.line_num == 1 and self.input_first_line_header:
                    self.input_header = row
                else:
                    self.input_data.append(row)
            input_file.close()

    def write_output_data_to_csv(self, output_filename):
        """Write output data in CSV file.

        :param output_filename: Path of the output_filename.
        """
        with open(output_filename, 'w', encoding=self.output_encoding, newline = '') as output_file:
            writer = csv.writer(output_file, delimiter = self.output_csv_delimiter)
            for row in self.output_data:
                writer.writerow(row)
            output_file.close()

    def read_input_data_from_xls(self, input_filename):
        """Read input data in old Excel file (xls)

        :param input_filename: Path of the input file
        """
        self.input_header = []
        self.input_data = []

        if not os.path.exists(input_filename):
            raise Exception('Error: Input file not found.')

        workbook = xlrd.open_workbook(filename = input_filename)
        worksheet = None
        if self.input_xls_sheet_name is None:
            worksheet = workbook.sheet_by_index(0)
        else:
            worksheet = workbook.sheet_by_name(self.input_xls_sheet_name)
        
        for row_index, row_data in enumerate(range(worksheet.nrows)):
            if row_index == 0 and self.input_first_line_header:
                for column_index, column_data in enumerate(range(worksheet.ncols)):
                    self.input_header.append(worksheet.cell_value(row_index, column_index))
            else:
                row = []
                for column_index, column_data in enumerate(range(worksheet.ncols)):
                    row.append(worksheet.cell_value(row_index, column_index))
                self.input_data.append(row)
        
    def write_output_data_to_xls(self, output_filename):
        """Write output data in old Excel file (xls)

        :param output_filename: Path of the output file
        """
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(self.output_xls_sheet_name)

        for row_index, row in enumerate(self.output_data):
            for col_index, col in enumerate(row):
                worksheet.write(row_index, col_index, col)
        workbook.save(output_filename)

    def read_input_data_from_xlsx(self, input_filename):
        """Read input data in new Excel file (xlsx)

        :param input_filename: Path of the input file
        """
        self.input_header = []
        self.input_data = []

        if not os.path.exists(input_filename):
            raise Exception('Error: Input file not found.')

        workbook = openpyxl.load_workbook(input_filename)
        worksheet = None
        if self.input_xls_sheet_name is None:
            worksheet = workbook.worksheets[0]
        else:
            worksheet = workbook.get_sheet_by_name(self.input_xls_sheet_name)

        for row_index, row_data in enumerate(worksheet.rows):
            if row_index == 0 and self.input_first_line_header:
                for col_data in row_data:
                    self.input_header.append(col_data.value)
            else:
                row = []
                for col_data in row_data:
                    row.append(col_data.value)
                self.input_data.append(row)
        
    def write_output_data_to_xlsx(self, output_filename):
        """Write output data in new Excel file (xlsx)

        :param output_filename: Path of the output file
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.output_xls_sheet_name

        for row in self.output_data:
            worksheet.append(row)
        workbook.save(output_filename)


    def convert(self):
        """Convert input data
        """
        self.output_data = []
        if self.input_first_line_header and not self.ignore_first_line_header:
            self.output_data.append(self.convert_row(self.input_header, 0, True))
        for row_index, row in enumerate(self.input_data):
            # Real row for index
            self.output_data.append(self.convert_row(row, row_index + 1))

    def convert_row(self, input_row, row_index, only_moves = False):
        """Convert row from input to output with moves and transform

        :param input_row: Input row data
        :param only_moves: Don't apply transforms

        :return: Converted row
        :rtype: list
        """
        output_row = [None] * self.output_number_of_cols
        if len(self.moves) > 0:
            for move in self.moves:
                input_col = input_row[int(move['from']) - 1]
                if 'action' in move and not only_moves:
                    input_col = self.action_input(move['action'], input_col, input_row, row_index)
                # transform if exists
                if 'transform' in move and not only_moves:
                    input_col = self.transform_input(move['transform'], input_col)
                output_row[int(move['to']) - 1] = input_col
        else:
            output_row = input_row
        return output_row

    def action_input(self, action_data, input_data, current_row, current_index):
        """Action on the column data.

        :param action_data: Action information (can be an array of actions).
        :param input_data: Input column.
        :param current_row: List of the columns of the current row.
        :param current_index: Index of the current row.

        :return: Column data after action.
        """
        if isinstance(action_data, list):
            for action_item in action_data:
                input_data = self.action_input(action_item, input_data, current_row, current_index)
        else:
            action_name = next(iter(action_data))
            input_data =  self.dynamics_modules[action_name].action(input_data, action_data[action_name], current_row, current_index)
        return input_data
        
    def transform_input(self, transform_data, input_data):
        """Transform the column data

        :param transform_data: Transform information (can be an array of actions).
        :param input_data: Input column.

        :return: Column data after transformation.
        """
        if isinstance(transform_data, list):
            for transform_item in transform_data:
                input_data = self.transform_input(transform_item, input_data)
        else:
            input_data =  self.dynamics_modules[transform_data].transform(input_data)
        return input_data
        
    def start(self, input_filename, output_filename, config = None):
        """Start conversion.

        :param input_filename: Name of the input file.
        :param output_filename: Name of the output file.
        :param config_file: Config file for conversion (optionnal).
        """

        self.set_default_values()

        # Read onvert informations
        if config is not None:
            self.read_config(config)

        # Test files types
        if self.input_file_type == self.UNKNOWN_FILE:
            self.input_file_type = self.get_file_type_by_ext(input_filename)
        if self.output_file_type == self.UNKNOWN_FILE:
            self.output_file_type = self.get_file_type_by_ext(output_filename)

        # Test if all files are good
        if (self.input_file_type == self.UNKNOWN_FILE or
            self.output_file_type == self.UNKNOWN_FILE):
            raise Exception('Error: Unknow file type.')

        # Test if input file exists
        if not os.path.exists(input_filename):
            raise Exception('Error: Input file missing')
        
        # Read data
        if self.input_file_type == self.CSV_FILE:
            self.read_input_data_from_csv(input_filename)
        elif self.input_file_type == self.OLD_EXCEL_FILE:
            self.read_input_data_from_xls(input_filename)
        elif self.input_file_type == self.NEW_EXCEL_FILE:
            self.read_input_data_from_xlsx(input_filename)

        # Convert
        self.convert()
            
        # Write data
        if self.output_file_type == self.CSV_FILE:
            self.write_output_data_to_csv(output_filename)
        elif self.output_file_type == self.OLD_EXCEL_FILE:
            self.write_output_data_to_xls(output_filename)
        elif self.output_file_type == self.NEW_EXCEL_FILE:
            self.write_output_data_to_xlsx(output_filename)

def usage(exec_name):
    """Show usage when this script is called from command line without enough arguments
    :param exec_name: Path of this script
    """

    # Get only executable name without the path
    if os.path.sep in exec_name:
        exec_name = exec_name.split(os.path.sep)[-1:][0]
        
    print(exec_name+" input_file output_file config")
    print(" - Input and output can be xls, xlsx, csv")
    print(" - Config can be a json file or json string")
    
# Entry point
if __name__ == '__main__':
    convert = Convert()

    argv = sys.argv
    if len(argv) < 3:
        usage(argv[0])
    elif len(argv) == 3:
        convert.start(input_filename = argv[1], output_filename = argv[2])
    elif len(argv) >= 4:
        convert.start(input_filename = argv[1], output_filename = argv[2], config = argv[3])
